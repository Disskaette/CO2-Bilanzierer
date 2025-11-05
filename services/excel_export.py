"""
Excel Export Service für CO₂-Bilanzierung
"""

import logging
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

from models.project import Project

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel Export für CO₂-Bilanzierung"""

    def export(
        self,
        filepath: str,
        project: Project,
        include_charts: bool = False
    ) -> bool:
        """Exportiert Excel mit allen Tabellen"""
        try:
            logger.info(f"Starte Excel-Export: {filepath}")

            wb = Workbook()
            wb.remove(wb.active)

            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F6AA5", end_color="1F6AA5", fill_type="solid")
            sum_font = Font(bold=True, size=11)
            sum_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Dashboard-Sheet
            ws_summary = wb.create_sheet("Dashboard")
            ws_summary.append(['Projektname:', project.name])
            ws_summary.append(['Systemgrenze:', project.system_boundary])
            ws_summary.append(['Erstellt am:', datetime.now().strftime('%d.%m.%Y %H:%M')])
            ws_summary.append([])

            ws_summary.append(['Variante', 'CO₂-Gesamt [t]'])
            
            for variant in project.variants:
                value = self._get_variant_total(variant, project.system_boundary)
                ws_summary.append([variant.name, value / 1000.0])

            # Header formatieren
            for cell in ws_summary[5]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Daten formatieren
            for row in ws_summary.iter_rows(min_row=6, max_row=5+len(project.variants)):
                for cell in row:
                    cell.border = thin_border
                    if cell.column == 2:
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right')

            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 15

            # Varianten-Sheets
            for idx, variant in enumerate(project.variants):
                ws = wb.create_sheet(f"{variant.name}")

                ws.append(['Pos', 'Material', 'Menge', 'Einheit', 'GWP A1-A3', 
                          'GWP C3', 'GWP C4', 'GWP D', 'CO₂-Gesamt [t]'])

                for i, row in enumerate(variant.rows, 1):
                    value = self._get_value_for_boundary(row, project.system_boundary)
                    ws.append([
                        i,
                        row.material_name,
                        row.quantity,
                        row.material_unit,
                        row.material_gwp_a1a3,
                        row.material_gwp_c3,
                        row.material_gwp_c4,
                        row.material_gwp_d if row.material_gwp_d else '',
                        value / 1000.0
                    ])

                total = self._get_variant_total(variant, project.system_boundary)
                sum_row = ['', 'SUMME', '', '', '', '', '', '', total / 1000.0]
                ws.append(sum_row)

                # Header
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                # Daten
                for row_idx in range(2, ws.max_row):
                    for cell in ws[row_idx]:
                        cell.border = thin_border
                        if cell.column == 1:
                            cell.alignment = Alignment(horizontal='center')
                        elif cell.column >= 3:
                            cell.number_format = '0.00'
                            cell.alignment = Alignment(horizontal='right')

                # Summenzeile
                for cell in ws[ws.max_row]:
                    cell.font = sum_font
                    cell.fill = sum_fill
                    cell.border = thin_border
                    if cell.column >= 3:
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='right')

                # Spaltenbreiten
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 35
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 12
                ws.column_dimensions['H'].width = 12
                ws.column_dimensions['I'].width = 15

                # Optional: Diagramm
                if include_charts and len(variant.rows) > 0:
                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    chart.title = f"{variant.name} - CO₂-Bilanz"
                    chart.y_axis.title = 'CO₂ [t]'
                    chart.x_axis.title = 'Material'

                    data = Reference(ws, min_col=9, min_row=1, max_row=len(variant.rows)+1)
                    cats = Reference(ws, min_col=2, min_row=2, max_row=len(variant.rows)+1)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                    
                    ws.add_chart(chart, f"K2")

            wb.save(filepath)
            logger.info("Excel erfolgreich erstellt")
            return True

        except Exception as e:
            logger.error(f"Excel-Export Fehler: {e}", exc_info=True)
            return False

    def _get_value_for_boundary(self, row, boundary: str) -> float:
        """Holt korrekten Wert basierend auf Systemgrenze"""
        if 'bio' in boundary.lower():
            if 'D' in boundary:
                return row.result_acd_bio or row.result_acd or 0.0
            elif 'C3' in boundary or 'C4' in boundary:
                return row.result_ac_bio or row.result_ac
            else:
                return row.result_a_bio or row.result_a
        else:
            if 'D' in boundary:
                return row.result_acd or 0.0
            elif 'C3' in boundary or 'C4' in boundary:
                return row.result_ac
            else:
                return row.result_a

    def _get_variant_total(self, variant, boundary: str) -> float:
        """Holt Variantensumme basierend auf Systemgrenze"""
        if 'bio' in boundary.lower():
            if 'D' in boundary:
                return variant.sum_acd_bio or variant.sum_acd or 0.0
            elif 'C3' in boundary or 'C4' in boundary:
                return variant.sum_ac_bio or variant.sum_ac
            else:
                return variant.sum_a_bio or variant.sum_a
        else:
            if 'D' in boundary:
                return variant.sum_acd or 0.0
            elif 'C3' in boundary or 'C4' in boundary:
                return variant.sum_ac
            else:
                return variant.sum_a
